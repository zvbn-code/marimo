import marimo

__generated_with = "0.23.2"
app = marimo.App(width="medium", sql_output="pandas")


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    # Auswerten der RBL-Aufzeichnungen aus IVU.control
    """)
    return


@app.cell
def _():
    import marimo as mo
    import duckdb
    import pandas as pd
    import altair as alt
    import datetime as dt 
    from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, Protection, PatternFill

    return Font, alt, dt, duckdb, mo, pd


@app.cell
def _(dt, mo):
    start_datum = mo.ui.date(label="Start Datum", value= '2025-10-01')
    ende_datum = mo.ui.date(label="Ende Datum", value= dt.datetime.now().strftime('%Y-%m-%d'))
    return ende_datum, start_datum


@app.cell
def _(ende_datum, mo, start_datum):
    mo.vstack([start_datum, ende_datum, mo.md(f"Zeitbereich Auswertung (inkl.): {start_datum.value} bis {ende_datum.value}")])
    return


@app.cell(hide_code=True)
def _(duckdb):
    db = 'ivu_rt_superset.db' # aus dem Verzeichnis python/ivu/db
    duck = duckdb.connect(db, read_only=False)
    return (duck,)


@app.cell(hide_code=True)
def _(duck, mo, rt_red):
    _df = mo.sql(
        f"""
        describe rt_red
        """,
        engine=duck
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Erstellen einer Kalenderdatei
    """)
    return


@app.cell(hide_code=True)
def _(duck, mo):
    _df = mo.sql(
        f"""
        create or replace table cal as
        select
            *
        from
            "https://daten.zvbn.de/all_tg.csv"
        """,
        engine=duck
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Auswertung der gemessenen Fahrten in dem Zeitraum
    """)
    return


app._unparsable_cell(
    r"""
    select
        cal.datum.strftime ('%Y-%m-%d') as datum,
        count(*) as anzahl_zeilen,
        count(*) filter(abwan not null) an_zeilen_ohne_null,
        count(distinct linie) as anzahl_linien,
    from
        cal
        left join rt_red on cal.datum = rt_red.datum
    where
        cal.datum >= '{start_datum.value}'
        and cal.datum <= '{ende_datum.value}'
    group by all
    order by all
    """,
    column=None, disabled=False, hide_code=True, name="_"
)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Testen der Ersetzenfunktion wegen unterschiedlicher Codierung
    """)
    return


@app.cell(hide_code=True)
def _(duck, mo, rt_red):
    _df = mo.sql(
        f"""
        select
            haltestelle_name,
            haltestelle_name.replace ('Ã¼', 'ü').replace ('Ã', 'ß').replace ('Ã¶', 'ö').replace ('Ã¤', 'ä').replace ('ß¤', 'ä') as hst_convert
        from
            rt_red
        where
            hst_convert like '%Nordwohlde%'
        group by all
        """,
        engine=duck
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Erstellen der Funktionen
    - df_res Erstellen des DataFrames für die ausgewählte Fahrt
    - chart_func Erstellung des Charts unterschieden nach html und pdf
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Funktionen Datenauswertung
    """)
    return


@app.cell
def _(duck):
    def df_res(fnr, start, ende):
        """Erstellt den Ergebnisdataframe für die ausgewählte Fahrt
        Rausfiltern besonders hoher Abweichungen und Begrenzung des Zeitbereichs
        """
        df = duck.sql(f"""
        select
        datum,
        kurs,
        mod(kurs, 2) as ri,
        linie,
        nr,
        haltestelle_name,
        sollab_ts,
        -- Ersetzen der Zeichen wegen unterschiedlicher Codierung der Haltestellennamen
        lpad(nr::TEXT, 2, '0') || ' ' || haltestelle_name.replace ('Ã¼', 'ü').replace ('Ã', 'ß').replace ('Ã¶', 'ö').replace('Ã¤', 'ä').replace('ß¤', 'ä') as nr_name,
        abwab / 60 as ab_minute,
        abwan / 60 as an_minute
        from
            rt_red
        where
            kurs = {fnr}
            -- Bestimmen des Toleranzbereichs für Abweichungen nach oben und unten 
            and abwab < 1800
            and abwab > -120
            and abwan < 1800
            and abwan > -120
            -- Zeitbereich
            and datum >= '{start}'
            and datum <= '{ende}'
            """).df()
        return df

    return (df_res,)


@app.cell
def _(duck):
    def df_res_list(fnr):
        """Erstellt den Ergebnisdataframe für die ausgewählte Fahrt
        Rausfiltern besonders hoher Abweichungen und Begrenzung des Zeitbereichs über Liste der verkehrenden Tage
        """
        df = duck.sql(f"""
        select
            l.datum,
            l.kurs,
            mod(l.kurs, 2) as ri,
            l.linie,
            l.buendel,
            l.nr,
            l.haltestelle_name,
            l.sollabfahrt,
            l.sollab_ts,
            -- Ersetzen der Zeichen wegen unterschiedlicher Codierung der Haltestellennamen
            lpad(l.nr::TEXT, 2, '0') || ' ' || l.haltestelle_name.replace ('Ã¼', 'ü').replace ('Ã', 'ß').replace ('Ã¶', 'ö').replace ('Ã¤', 'ä').replace ('ß¤', 'ä') as nr_name,
            l.abwab / 60 as ab_minute,
            l.abwan / 60 as an_minute
        from
            rt_red l
            join tbl_fahrtliste s on s.kurs = l.kurs
            and l.datum in s.list_datum
        where
            l.kurs = {fnr}
            and abwab < 1800
            and abwab > -120
            and abwan < 1800
            and abwan > -120
        order by datum, kurs, nr
        """).df()
        return df

    return (df_res_list,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Funktion Erstellung Chart
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    #### Chart html
    """)
    return


@app.cell
def _(alt):
    def chart_func(df_in, x,y, title, suptitle):
        """
        Erstellt aus einem Dateframe je Fahrt die Ausgabe als html mit Nutzung der gesamten Breite (container)
        """
        chart = alt.Chart(
            df_in, 
            title=alt.TitleParams(text=title, anchor='start', fontSize=14, subtitle=suptitle, subtitleFontSize=12), 
            width='container', 
            height=500
        ).mark_boxplot(
            extent=1.5, 
            box = {'color': 'lightblue'},
            median={'color': 'orangered'},
            outliers={'size':3, 'color':'darkgrey', 'fill':'darkgrey', 'tooltip':{'content':'data'} },
            ticks={'color':'green', 'size':8}
        ).encode(
            alt.X(f"{x}:N").title('lfd. Nr. / Haltestelle'),
            alt.Y(f"{y}:Q").scale(zero=False).title('Abweichung Ankunft in Minuten'),  
            #tooltip = f"{x}"
        ).configure_title(
            fontWeight='normal', fontSize=12
        ).configure_axis(
            labelFontWeight='normal',
            titleFontSize = 14,
            titleFontWeight='normal', 
            labelFontSize = 12
        )
        return chart


    return (chart_func,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    #### Chart pdf
    """)
    return


@app.cell
def _(alt):
    def chart_func_pdf(df_in, x,y, title, suptitle):
        """
        Erstellt aus einem Dateframe je Fahrt die Ausgabe als pdf mit Nutzung mit einem festen Seitenverhälntis
        """
        chart = alt.Chart(
            df_in, 
            title=alt.TitleParams(text=title, anchor='start', fontSize=14, subtitle=suptitle, subtitleFontSize=12), 
            width=1000, 
            height=500
        ).mark_boxplot(
            extent=1.5, 
            box = {'color': 'lightblue'},
            median={'color': 'orangered'},
            outliers={'size':3, 'color':'darkgrey', 'fill':'darkgrey'},
            ticks={'color':'green', 'size':8}
        ).encode(
            alt.X(f"{x}:N").title('lfd. Nr. / Haltestelle'),
            alt.Y(f"{y}:Q").scale(zero=False).title('Abweichung Ankunft in Minuten'),  
        ).configure_title(
            fontWeight='normal', fontSize=12
        ).configure_axis(
            labelFontWeight='normal',
            titleFontSize = 14,
            titleFontWeight='normal', 
            labelFontSize = 12
        )
        return chart

    return (chart_func_pdf,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Erstellen der Fahrtliste und Ermitteln der plausiblen Fahrten
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Ermitteln der Häufigkeit einer Fahrt auch bei unterschiedlichen Anzahlen von Haltestellen
    """)
    return


@app.cell(hide_code=True)
def _(duck, ende_datum, mo, rt_red, start_datum):
    _df = mo.sql(
        f"""
        -- erstellen einer temporären Tabelle mit der Liste der Fahrten und auswertbaren Tagen
        create or replace temporary table tbl_fahrtliste as
        select
            * exclude (anz)
        from
            (
                select
                    buendel,
                    linie,
                    anz_hst,
                    kurs,
                    mod(kurs,2) as ri,
                    anz,
                    sum(anz) over (
                        partition by
                            kurs
                    ) as ges_erhoben,
                    max(anz) over (
                        partition by
                            kurs
                    ) as anz_erhoben,
                    list_datum,
                    -- list_datum_sep,
                    -- list_string_agg(list_datum) as list_datum_str,
                    list_min(list_datum) as von,
                    list_max(list_datum) as bis,
                    '<a href="chart/chart_' || kurs::int || '.html">Link</a>' as link
                from
                    (
                        select
                            buendel,
                            linie,
                            kurs,
                            anz_hst,
                            count(*) as anz,
                            -- STRING_AGG(datum.strftime ('%Y-%m-%d')::text, ',') as list_datum_sep,
                            list(datum) as list_datum
                        FROM
                            (
                                select
                                    datum,
                                    buendel,
                                    linie,
                                    kurs,
                                    count(*) as anz_hst,
                                    count(*) filter(abwan not null) as an_not_null,
                                    count(*) filter(abwab not null) as ab_not_null,
                                from
                                    rt_red
                                where
                                    datum >= '{start_datum.value}'
                                    and datum <= '{ende_datum.value}'
                                    -- and linie = 6440
                                    -- and abwab < 1800
                                    -- and abwab > -120
                                    -- and abwan < 1800
                                    -- and abwan > -120
                                group by all
                                order by
                                    kurs,
                                    datum
                            ) as foo
                        where
                            ab_not_null > 2
                            and an_not_null > 2
                            and (ab_not_null / anz_hst) > 0.75
                            and (an_not_null / anz_hst) > 0.75
                        group by all
                    ) as foo2
                order by
                    kurs
            )
        WHERE
            -- Rausfiltern der Fahrt die am öftestesten erhoben wurde
            anz = anz_erhoben
            and anz >= 4
        order by
            kurs
        """,
        engine=duck
    )
    return


@app.cell
def _(df_res_list):
    df_res_list(fnr=1226001)
    return


@app.cell
def _(duck):
    # Filterung auf Fahrten 
    ## - mit einer gewissen Anzahl von Fahrten und 
    ## - gleicher Anzahl Haltestellen und 
    ## - mehr als zwei Haltestellen mit fehlender Angaben zu Abwwichung An/ab
    df_fahrten_sel = duck.sql("from tbl_fahrtliste").df()
    return (df_fahrten_sel,)


@app.cell(hide_code=True)
def _(duck, mo, rt_red, start_datum):
    _df = mo.sql(
        f"""
        select datum, kurs, max(abwab) 
            from rt_red 
            where linie = 6225 and datum >= '{start_datum.value}' 
            group by all 
            order by kurs, datum
        """,
        engine=duck
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Erstellen der Liste gesamt und je Bündel
    """)
    return


@app.cell
def _(df_fahrten_sel):
    df_fahrten_sel[['buendel','linie', 'kurs','ri' ,'von', 'bis', 'anz_erhoben', 'link']].style.format({"von": lambda t: t.strftime("%Y-%m-%d"), "bis": lambda t: t.strftime("%Y-%m-%d")}, precision=0).to_html('out/liste.html', index=False, escape=False)
    return


@app.cell(hide_code=True)
def _(duck, mo, tbl_fahrtliste):
    df_list_buendel = mo.sql(
        f"""
        select distinct
            buendel
        from
            tbl_fahrtliste
            where buendel not null
        order by
            buendel
        """,
        engine=duck
    )
    return (df_list_buendel,)


@app.cell
def _(df_fahrten_sel, df_list_buendel):
    for j, value in df_list_buendel.iterrows():
        _buendel = value['buendel']
        print(_buendel.replace(' ', '_').replace('ü', 'ue').lower())
        df_fahrten_sel[['buendel','linie', 'kurs', 'von', 'bis', 'anz_erhoben', 'link']].query(f"buendel == '{_buendel}'").style.format({"von": lambda t: t.strftime("%Y-%m-%d"), "bis": lambda t: t.strftime("%Y-%m-%d")}, precision=0).to_html(f"out/liste_{_buendel.replace(' ', '_').replace('ü', 'ue').lower()}.html", index=False, escape=False)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Batchdurchlauf
    """)
    return


@app.cell
def _(df_res_list):
    df_res_list(fnr=1226002)
    return


@app.cell
def _(df_res_list, pd):
    # Ermitteln der Median Werte 
    _arr = []
    _fnr = 1226005
    _df = df_res_list(fnr=_fnr)
    _median_erste = _df.query("nr == 1")['ab_minute'].median()
    _vorletzte_hst = _df.nr.max()-1
    _median_vorletzte = _df.query(f"nr == {_vorletzte_hst}")['an_minute'].median()
    _buendel = _df['buendel'].drop_duplicates()[0]
    _ri = _df['ri'].drop_duplicates()[0]
    _linie = _df['linie'].drop_duplicates()[0]
    _hst_ab = _df.query("nr == 1")['haltestelle_name'].min()
    _arr.append([_fnr,_buendel,_linie,_median_erste, _median_vorletzte, _hst_ab, _ri])
    pd.DataFrame(_arr, columns=['fnr','buendel','linie','median_erste', 'median_vorletzte', 'hst_ab', 'ri'])
    #df_res_list(fnr=1226002)
    return


@app.cell
def _(df_fahrten_sel):
    df_fahrten_sel.to_excel('df_fahrten_sel.xlsx')
    sel_parquet = 'df_fahrten_sel.parquet'
    df_fahrten_sel.to_parquet(sel_parquet)
    df_fahrten_sel
    return (sel_parquet,)


@app.cell
def _(df_fahrten_sel):
    df_fahrten_sel.query('linie in ("1330", "1340")')
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Abgleich der Fahrten mit dem Kalender
    """)
    return


@app.cell(hide_code=True)
def _(mo, sel_parquet):
    _df = mo.sql(
        f"""
        select * from read_parquet({sel_parquet})
        """
    )
    return


@app.cell
def _():
    return


@app.cell(hide_code=True)
def _(cal, duck, ende_datum, mo, sel_parquet, start_datum):
    _df = mo.sql(
        f"""
        pivot (
            select
                cal.datum,
                cal.tg1,
                cal.tg2,
                f.linie,
                f.kurs
            from
                cal
                left join (
                    from
                        read_parquet({sel_parquet})
                ) f on cal.datum in f.list_datum
            where
                datum >= '{start_datum.value}'
                and datum <= '{ende_datum.value}'
            and linie in (1330, 1340,1630)
        ) on linie using count(kurs)
        group by
            datum
        """,
        engine=duck
    )
    return


@app.cell
def _(df_fahrten_sel):
    print(len(df_fahrten_sel.query('linie in ("1330", "1340")')))
    return


@app.cell
def _(chart_func, chart_func_pdf, df_fahrten_sel, df_res_list, dt, pd):
    _arr = []

    for _k, _value in df_fahrten_sel.query('linie in ("1330", "1340")')[1:5].iterrows():
        _fnr = int(_value['kurs'])
        _df = df_res_list(fnr=_fnr)

        _min_datum = _df.datum.min().date().strftime('%Y-%m-%d')
        _max_datum = _df.datum.max().date().strftime('%Y-%m-%d')    
        _anzahl = _df.datum.drop_duplicates().count()    

        # Ermitteln der Median Werte je Fahrt erste ab und vorletzte an
        _median_erste = _df.query("nr == 1")['ab_minute'].median()
        _sollab = _df.query("nr == 1")['sollab_ts'].min()
        _list_datum = _df['datum'].drop_duplicates().sort_values().tolist()
        vorletzte_hst = _df.nr.max()-1
        _median_vorletzte = _df.query(f"nr == {vorletzte_hst}")['an_minute'].median()
        _buendel = _df['buendel'].drop_duplicates()[0]
        _linie = _df['linie'].drop_duplicates()[0]
        _ri = _df['ri'].drop_duplicates()[0]
        _hst_ab = _df.query("nr == 1")['haltestelle_name'].min()
        _arr.append([_fnr,_buendel,_linie,_median_erste, _median_vorletzte, _min_datum, _max_datum, _anzahl, _sollab, _sollab.hour, _sollab.minute, _hst_ab, _ri])

        print(_fnr, _min_datum, _max_datum, _anzahl, _sollab.hour, _sollab.minute)

        #Erstellen des Charttitels und Untertitels mit den wichtigsten Informationen
        _title = f"Fahrt {_fnr} von {_min_datum} bis {_max_datum} Anzahl {_anzahl} Start {_hst_ab} {_sollab.hour}:{str(_sollab.minute).zfill(2)}"
        _erstellt = dt.datetime.now().strftime('%Y-%m-%d %H:%M')
        _list =_df.datum.drop_duplicates().dt.strftime('%Y-%m-%d').sort_values().tolist()
        _len_list = len(_list)
        _str_list = ', '.join(_list[0:5])
        if _len_list > 5:
            _str_list += f", ... ({_len_list -5} weitere Tage)"

        _chart = chart_func(df_in=_df, x='nr_name', y='an_minute', title=_title, suptitle = f"Erstellt: {_erstellt} Erhobene Tage {_str_list}")
        _chart_pdf = chart_func_pdf(df_in=_df, x='nr_name', y='an_minute', title=_title, suptitle = f"Erstellt: {_erstellt} Erhobene Tage {_str_list}")
        _chart_pdf.save(f'out/chart/chart_{_fnr}.pdf')
        _chart.save(f'out/chart/chart_{_fnr}.html')
        #print(_list_datum)

    df_median = pd.DataFrame(_arr, columns=['fnr','buendel','linie','median_erste', 'median_vorletzte', 'min_datum', 'max_datum', 'anzahl', 'sollab', 'stunde', 'minute', 'Hst ab', 'ri'])
    return (df_median,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Ausgabe Medianberichte
    """)
    return


@app.cell
def _(Font, df_median, ende_datum, pd, start_datum):
    prefix = f"reports/median_{start_datum.value} - {ende_datum.value}"

    with pd.ExcelWriter(f'{prefix}.xlsx', engine='openpyxl') as writer:    
        df_median.sort_values(['buendel', 'linie', 'fnr'], ascending=True).style.format(precision=1).to_excel(writer, sheet_name='median', index=False)
        workbook  = writer.book
        ws = writer.sheets['median']
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'
        ws.column_dimensions['F'].width = 15 # Datum
        ws.column_dimensions['G'].width = 15 # Datum
        ws.column_dimensions['L'].width = 25 # Anzahl


        for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
            for _cell in row:
                _cell.number_format = "0.00" # Display to 2dp
                if type(_cell.value) in [int, float] and (_cell.value > 5  or _cell.value < -1):
                    _cell.font = Font(color="00FF0000")

    df_median.to_parquet(f'{prefix}.parquet')
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Grafische Darstellung der Ergebnisse
    - Auswertung der Medianwerte nach Uhrzeit
    """)
    return


@app.cell
def _(df_median, mo):
    dropdown = mo.ui.dropdown.from_series(df_median["linie"])
    return (dropdown,)


@app.cell
def _(dropdown):
    dropdown
    return


@app.cell(hide_code=True)
def _(alt, df_median, dropdown):
    _linie = 6440
    _df = df_median[df_median['sollab'].notnull()].query(f"linie == '{dropdown.value}'").sort_values(['ri', 'stunde', 'fnr']).reset_index()

    _chart = alt.Chart(_df[['fnr','stunde','minute' ,'median_vorletzte', 'ri']],  title=alt.TitleParams(text=f"Fahrtzeiten Ankunft vorletzte Haltestelle Linie {dropdown.value}", anchor='start', fontSize=14)).mark_circle(size=60).encode(
        alt.X("fnr:N"),
        alt.Y("stunde:N"),
        alt.Color("median_vorletzte:Q", scale=alt.Scale(scheme='redblue', reverse=True, domainMid=5)),
        tooltip=['fnr', 'stunde','minute' ,'median_vorletzte', 'ri'],


    )
    chart_agg = _chart
    return (chart_agg,)


@app.cell
def _(chart_agg):
    chart_agg
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Einzeldurchlauf
    - nur für einen Zeitraum von bis , hier müssen die Verläufe identisch sein
    """)
    return


@app.cell
def _(df_res, ende_datum, start_datum):
    _list =df_res(fnr=6225080, start=start_datum.value, ende=ende_datum.value).datum.drop_duplicates().dt.strftime('%Y-%m-%d').sort_values().tolist()
    _str_list = ', '.join(_list)
    print(_str_list)
    return


@app.cell
def _(chart_func, df_res, dt, ende_datum, start_datum):
    _fnr = 6225080
    _df = df_res(fnr=_fnr, start=start_datum.value, ende=ende_datum.value)
    _min_datum = _df.datum.min().date().strftime('%Y-%m-%d')
    _max_datum = _df.datum.max().date().strftime('%Y-%m-%d')    
    _anzahl = _df.datum.drop_duplicates().count() 

    # Ermitteln der Median Werte je Fahrt erste ab und vorletzte an

    _sollab = _df.query("nr == 1")['sollab_ts'].min()
    _list_datum = _df['datum'].drop_duplicates().sort_values().tolist()
    _hst_ab = _df.query("nr == 1")['haltestelle_name'].min()

    #Erstellen des Charttitels und Untertitels mit den wichtigsten Informationen
    _title = f"Fahrt {_fnr} von {_min_datum} bis {_max_datum} Anzahl {_anzahl} Start {_hst_ab} {_sollab.hour}:{str(_sollab.minute).zfill(2)}"
    _erstellt = dt.datetime.now().strftime('%Y-%m-%d %H:%M')
    _list =_df.datum.drop_duplicates().dt.strftime('%Y-%m-%d').sort_values().tolist()
    _len_list = len(_list)
    _str_list = ', '.join(_list[0:5])
    if _len_list > 5:
        _str_list += f", ... ({_len_list -5} weitere Tage)"

    _chart = chart_func(df_in=_df, x='nr_name', y='an_minute', title=_title, suptitle = f"Erstellt: {_erstellt} Erhobene Tage {_str_list}")
    _chart
    _chart.save(f'out/chart_{_fnr}.pdf')
    _chart.save(f'out/chart_{_fnr}.html')
    chart_intern = _chart
    return (chart_intern,)


@app.cell
def _(chart_intern):
    chart_intern
    return


@app.cell
def _(duck):
    duck.close()
    return


if __name__ == "__main__":
    app.run()
